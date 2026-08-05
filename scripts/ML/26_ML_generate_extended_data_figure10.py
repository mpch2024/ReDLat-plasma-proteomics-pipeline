###############################################################################
# ReDLat plasma proteomics — strict machine-learning workflow
# 26. Generate matched-sensitivity Extended Data Figure 10
# Requires: Script 16 aggregate matched outputs
# Produces: publication figure and identifier-free Source Data workbook
# Data policy: participant-level OOF predictions are not exported.
###############################################################################

from pathlib import Path
import sys
_REPO_HINT = Path(__file__).resolve()
for _candidate in [_REPO_HINT.parent, *_REPO_HINT.parents]:
    if (_candidate / ".redlat-root").exists() or (_candidate / ".here").exists():
        PROJECT_ROOT = _candidate
        break
else:
    raise FileNotFoundError("Repository root not found. Set REDLAT_PROJECT_ROOT.")
sys.path.insert(0, str(PROJECT_ROOT / "python"))

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from redlat_ml.config import load_config, require_files, assert_public_table

CONFIG = load_config(__file__)
MATCHED = CONFIG.private_root / "matched"
RESELECT = MATCHED / "02_nested_cv"
FIXED = MATCHED / "03_primary_fixed_panel_cv"
PTAU = MATCHED / "04_primary_panel_ptau"
OUT = CONFIG.publication_root / "extended_data_figure_10"
OUT.mkdir(parents=True, exist_ok=True)

required = [
    (RESELECT / "metrics_nestedCV.csv", "matched reselected metrics"),
    (RESELECT / "outer_panel_frequency.csv", "matched selection frequency"),
    (FIXED / "metrics_nestedCV.csv", "matched fixed-panel metrics"),
    (FIXED / "permutation_importance.csv", "matched fixed-panel importance"),
    (PTAU / "metrics_all_models.csv", "matched p-tau217 metrics"),
    (PTAU / "roc_curves.csv", "matched p-tau217 ROC"),
    (PTAU / "delong_results.csv", "matched p-tau217 DeLong"),
    (PTAU / "cohort_summary.csv", "matched p-tau217 cohort summary"),
]
require_files(required)
reselect = pd.read_csv(RESELECT / "metrics_nestedCV.csv")
fixed = pd.read_csv(FIXED / "metrics_nestedCV.csv")
frequency = pd.read_csv(RESELECT / "outer_panel_frequency.csv")
importance = pd.read_csv(FIXED / "permutation_importance.csv")
ptau_metrics = pd.read_csv(PTAU / "metrics_all_models.csv")
roc = pd.read_csv(PTAU / "roc_curves.csv")
delong = pd.read_csv(PTAU / "delong_results.csv")
cohort = pd.read_csv(PTAU / "cohort_summary.csv")
for label, table in [("reselected", reselect), ("fixed", fixed), ("frequency", frequency), ("importance", importance), ("ptau_metrics", ptau_metrics), ("roc", roc), ("delong", delong), ("cohort", cohort)]:
    assert_public_table(table, label)

plt.rcParams.update({"font.family": "Arial", "font.size": 7, "pdf.fonttype": 42, "ps.fonttype": 42})
fig, axes = plt.subplots(2, 2, figsize=(180 / 25.4, 150 / 25.4))
ax = axes[0, 0]
summary = pd.DataFrame({
    "Model": ["Reselected panel", "Primary fixed panel"],
    "Mean_AUC": [reselect.AUC.mean(), fixed.AUC.mean()],
    "SD_AUC": [reselect.AUC.std(), fixed.AUC.std()],
})
ax.bar(summary.Model, summary.Mean_AUC, yerr=summary.SD_AUC, capsize=2)
ax.set_ylim(0.5, 1.0); ax.set_ylabel("ROC AUC"); ax.tick_params(axis="x", rotation=20); ax.set_title("Matched protein models")

ax = axes[0, 1]
grid = np.linspace(0, 1, 101)
for model, group in roc.groupby("Model"):
    curves=[]
    for _, fold in group.groupby("Fold"):
        curves.append(np.interp(grid, fold.FPR, fold.TPR))
    mean=np.mean(curves,axis=0); sd=np.std(curves,axis=0,ddof=1) if len(curves)>1 else np.zeros_like(mean)
    ax.plot(grid, mean, label=model); ax.fill_between(grid, np.clip(mean-sd,0,1), np.clip(mean+sd,0,1), alpha=.15)
ax.plot([0,1],[0,1],linestyle="--",linewidth=.6); ax.set_xlabel("False-positive rate"); ax.set_ylabel("True-positive rate"); ax.set_title("Matched biomarker comparison"); ax.legend(frameon=False, fontsize=6)

ax = axes[1, 0]
freq_col = "Selection_frequency" if "Selection_frequency" in frequency.columns else frequency.columns[-1]
feature_col = "Feature" if "Feature" in frequency.columns else frequency.columns[0]
plot_freq = frequency.sort_values(freq_col).tail(12)
ax.barh(plot_freq[feature_col], plot_freq[freq_col]); ax.set_xlabel("Outer-fold selection frequency"); ax.set_title("Matched panel stability")

ax = axes[1, 1]
plot_imp = importance.sort_values("Permutation").tail(12)
ax.barh(plot_imp.Proteins, plot_imp.Permutation, xerr=plot_imp.get("Permutation_sd", None)); ax.set_xlabel("Held-out permutation importance"); ax.set_title("Primary fixed panel")

for tag, ax in zip("abcd", axes.ravel()):
    ax.text(-0.14, 1.07, tag, transform=ax.transAxes, fontweight="bold", fontsize=9, va="top")
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
fig.tight_layout()
fig.savefig(OUT / "Extended_Data_Figure_10.pdf", bbox_inches="tight")
fig.savefig(OUT / "Extended_Data_Figure_10.svg", bbox_inches="tight")
fig.savefig(OUT / "Extended_Data_Figure_10.png", dpi=300, bbox_inches="tight")
plt.close(fig)

sheets = {
    "Panel_a_AUC": summary,
    "Panel_b_ROC": roc,
    "Panel_c_Frequency": frequency,
    "Panel_d_Importance": importance,
    "Matched_pTau_Metrics": ptau_metrics,
    "Matched_pTau_DeLong": delong,
    "Cohort_Summary": cohort,
}
wb=Workbook(); wb.remove(wb.active)
for name, table in sheets.items():
    ws=wb.create_sheet(name[:31]); ws.append(list(table.columns))
    for cell in ws[1]: cell.font=Font(bold=True); cell.fill=PatternFill("solid",fgColor="E9E5DC"); cell.alignment=Alignment(wrap_text=True)
    for row in table.itertuples(index=False, name=None): ws.append(list(row))
    ws.freeze_panes="A2"
wb.save(OUT / "Source_Data_Extended_Data_Figure_10.xlsx")
print(f"Generated: {OUT}")
